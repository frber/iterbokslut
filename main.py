from tkinter import *
from tkinter.ttk import *
#import tkinter as tk
#import tkinter.ttk as ttk
import pandas as pd
import openpyxl



def skapa_forslag(lista_box, lista_drop, lista_ent):
    wb = openpyxl.load_workbook('Docs/Proj.xlsx', data_only=True)
    ws = wb["Proj"]
    for box, drop, ent in zip(lista_box, lista_drop, lista_ent):
        if box[0].get() == 1:
            projnummer = box[1].split()[0]
            finansiar = drop.get()
            fingrad = ent.get()

            print(projnummer, finansiar, fingrad)



def skapa_box_drop(root, numnamn, finansiarer):
    lista_box = []
    lista_drop = []
    lista_ent = []
    rad = 1
    for x in numnamn:
        box = IntVar()
        drop = StringVar()
        input_text = StringVar()
        checkbox = Checkbutton(root, text=x, variable=box)
        checkbox.grid(row=rad, column=0, sticky="W")
        droplist = OptionMenu(root, drop, "", *finansiarer)
        droplist.grid(row=rad, column=1, sticky="W")
        ent = Entry(root, textvariable=input_text)
        ent.grid(row=rad, column=2, sticky="W")
        lista_box.append([box, x])
        lista_drop.append(drop)
        lista_ent.append(input_text)
        rad += 1

    knapp = Button(root, text="Skapa förslag", command=lambda: skapa_forslag(lista_box, lista_drop, lista_ent))
    knapp.grid(row=rad + 1, column=0)

def hamta_projekt():
    wb = openpyxl.load_workbook('Docs/Proj.xlsx', data_only=True)
    ws = wb["Proj"]
    projnummer = []
    projnamn = []
    numnamn = []
    for row in ws['A1:A100']:
        for cell in row:
            konto = cell.value
            proj = cell.offset(column = 2).value
            projt = cell.offset(column = 3).value
            belopp = cell.offset(column = 7).value
            if konto == None and proj != None and projt != None and belopp != None:
                projnummer.append(proj)
                projnamn.append(projt)
                numnamn.append(str(proj)+ " "+projt)
    wb.close()
    return projnummer, projnamn, numnamn

def hamta_fin():
    df = pd.read_excel(r'Docs\Data.xlsx')
    fin = df['FINANSIÄR'].tolist()
    return fin

def skapa_labels(root):
    rubfont = font = ('TkDefaultFont', 10, 'bold')
    projektlabel = Label(root, text="Projekt", font=rubfont)
    projektlabel.grid(row=0, column=0, sticky="s")
    finansiarlabel = Label(root, text="Finansiär", font=rubfont)
    finansiarlabel.grid(row=0, column=1, sticky="s")
    fingradlabel = Label(root, text="Finansieringsgrad", font=rubfont)
    fingradlabel.grid(row=0, column=2, sticky="s")

def main():
    root = Tk()
    root.geometry("680x350")

    skapa_labels(root)

    finansiarer = hamta_fin()
    projektnummer, projektnamn, numnamn = hamta_projekt()[0], hamta_projekt()[1], hamta_projekt()[2]
    skapa_box_drop(root, numnamn, finansiarer)

    root.mainloop()

if __name__ == "__main__":
    main()

#Ge dynamiskt periodiseringsförslag efter val av fin och fingrad?
#Sätta in förslag i ny eller tidgare använd berper (beroende på namn etc).
#Skapa rätt berpers med rätt namn utifrån kontering i berper
#Flytta över till bokföringsmall med/utan rev från berpers
#Kontrollera fel i berpers


#Lägga in data - ex kontrollera vht och göra färdigt en bokföringsorder med rättningsförslag. Kontrollera andra saker
#Visualisera kontering av personal

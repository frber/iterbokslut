from tkinter import *
from tkinter.ttk import *
from ttkthemes import ThemedTk
from tkinter import filedialog
import pandas as pd
import openpyxl
import os
import threading


from tab3 import *



class Tab2:
    def __init__(self, tab2, tab3):

        self.tab2 = tab2
        self.tab3 = tab3

        #ENTRY----------------------------------------------

        ent_y = 40
        ent_x = 40

        #  Namn
        self.finnamn_label = Label(self.tab2, text="Namn:")
        self.finnamn_label.place(y=ent_y, x=ent_x)
        self.finnamn = Entry(self.tab2, width=31)
        self.finnamn.place(y=ent_y, x=ent_x+45)

        # Motpart
        self.motp_label = Label(self.tab2, text="Motpart:")
        self.motp_label.place(y=ent_y+30, x=ent_x)
        self.motp = Entry(self.tab2, width=10)
        self.motp.place(y=ent_y+30, x=ent_x+55)

        # Perkonto fordran
        self.perf_label = Label(self.tab2, text="Perkonto fordran:")
        self.perf_label.place(y=ent_y+60, x=ent_x)
        self.perf = Entry(self.tab2, width=10)
        self.perf.place(y=ent_y+60, x=ent_x+100)

        # Perkonto skuld
        self.pers_label = Label(self.tab2, text="Perkonto skuld:")
        self.pers_label.place(y=ent_y+90, x=ent_x)
        self.pers = Entry(self.tab2, width=10)
        self.pers.place(y=ent_y+90, x=ent_x+100)


        # Radio OH
        radio_y = ent_y+120
        radio_x = 40
        self.ent_c = 0
        self.lista_ent = []
        self.lista_ent_label = []
        self.radio_var = IntVar()
        self.radio_ja = Radiobutton(self.tab2, text="Godkänner all OH", variable=self.radio_var, value=1, command=self.ta_bort_entry)
        self.radio_ja.place(y=radio_y, x=radio_x)
        self.radio_dir = Radiobutton(self.tab2, text="Godkänner % på totala direkta kostnader", variable=self.radio_var, value=2, command=self.skapa_entry)
        self.radio_dir.place(y=radio_y+30, x=radio_x)
        self.radio_dir = Radiobutton(self.tab2, text="Godkänner % på lönekostnader", variable=self.radio_var, value=3, command=self.skapa_entry)
        self.radio_dir.place(y=radio_y+60, x=radio_x)
        self.radio_ingen = Radiobutton(self.tab2, text="Godkänner ingen OH", variable=self.radio_var, value=4, command=self.ta_bort_entry)
        self.radio_ingen.place(y=radio_y+90, x=radio_x)

        # Listbox in
        self.lista_valda_kostnader = []
        self.kontogrupper = self.hamta_kontogrupper()
        self.listbox_in = Listbox(self.tab2, height='15', width='25', exportselection=False)
        self.listbox_in.place(y=50, x=400)

        self.listbox_in.bind('<<ListboxSelect>>', self.speca_kontogrupp)

        for a in self.kontogrupper:
            self.listbox_in.insert(END, a)
        #Label listbox in
        self.label_listbox_in = Label(self.tab2, text="Kontogrupper")
        self.label_listbox_in.place(y=20, x=430)

        # Listbox ut
        self.listbox_ut = Listbox(self.tab2, height='15', width='60')
        self.listbox_ut.place(y=50, x=700)
        # Label listbox ut
        self.listbox_ut_label = Label(self.tab2, text="Ej godkända kostnader")
        self.listbox_ut_label.place(y=20, x=810)

        # Listbox spec kontogrupp
        self.listbox_kontogrupp = Listbox(self.tab2, height='15', width='60')
        self.listbox_kontogrupp.place(y=50, x=1200)
        # Label listbox spec kontogrupp
        self.listbox_ut_label = Label(self.tab2, text="Kontospecifikation")
        self.listbox_ut_label.place(y=20, x=1330)




        # Knapp för över listbox kontogrupp
        self.knapp_lagg_till_grupp = Button(self.tab2, text="   --->", command=self.for_over_listbox_grupp)
        self.knapp_lagg_till_grupp.place(y=120, x=600)

        # Knapp ta bort listbox kontogrupp
        self.knapp_ta_bort_grupp = Button(self.tab2, text="   <---", command=self.ta_bort_listbox)
        self.knapp_ta_bort_grupp.place(y=190, x=600)

        # Knapp för över listbox kontospec
        self.knapp_lagg_till_spec = Button(self.tab2, text="   <---", command=self.for_over_listbox_spec)
        self.knapp_lagg_till_spec.place(y=120, x=1100)

        # Knapp ta bort listbox kontospec
        self.knapp_ta_bort_spec = Button(self.tab2, text="   --->", command=self.ta_bort_listbox)
        self.knapp_ta_bort_spec.place(y=190, x=1100)






        # Knapp Spara i db
        self.knapp_spara_db = Button(self.tab2, text="Spara till databas", command=self.spara_till_db)
        self.knapp_spara_db.place(y=630, x=150)

        # Knapp ta bort från db
        self.ta_bort = Button(self.tab2, text="Ta bort från databas", command=self.ta_bort_fran_db)
        self.ta_bort.place(y=630, x=1330)

        # Träd
        self.tree = Treeview(self.tab2)
        self.tree['columns'] = ("Namn", "Motpart", "Perkonto fordran", "Perkonto skuld", "OH", "Ej godkända kostnader")
        self.tree.column("#0", width=0, stretch=NO)
        self.tree.column("Namn", anchor=W)
        self.tree.column("Motpart", anchor=W, width=100)
        self.tree.column("Perkonto fordran", anchor=W, width=100)
        self.tree.column("Perkonto skuld", anchor=W, width=100)
        self.tree.column("OH", anchor=W)
        self.tree.column("Ej godkända kostnader", anchor=W, width=600)

        self.tree.heading("#0", text="", anchor=W)
        self.tree.heading("Namn", text="Namn", anchor=W)
        self.tree.heading("Motpart", text="Motpart", anchor=W)
        self.tree.heading("Perkonto fordran", text="Perkonto fordran", anchor=W)
        self.tree.heading("Perkonto skuld", text="Perkonto skuld", anchor=W)
        self.tree.heading("OH", text="OH", anchor=W)
        self.tree.heading("Ej godkända kostnader", text="Ej godkända kostnader", anchor=W)

        self.tree.place(y=400, x=150)
        self.initiera_trad()

    def hamta_kontogrupper(self):
        df = pd.read_excel(r'Docs\Konton.xlsx')
        kontogrupp = df['Rel.värde (T)'].tolist()
        kontogrupp = list(set(kontogrupp))
        return kontogrupp

    def skapa_entry(self):
        self.ent_c = 1
        self.oh_ent_label = Label(self.tab2, text="Ange godkänd %")
        self.oh_ent_label.place(y=280, x=40)
        self.oh_ent = Entry(self.tab2, width=10)
        self.oh_ent.place(y=280, x=150)

        self.lista_ent.append(self.oh_ent)
        self.lista_ent_label.append(self.oh_ent_label)

    def ta_bort_entry(self):
        if self.ent_c > 0:
            for x in self.lista_ent:
                x.destroy()
            for y in self.lista_ent_label:
                y.destroy()
        self.ent_c = 0

    def speca_kontogrupp(self, event):
        self.listbox_kontogrupp.delete(0, END)
        select = self.listbox_in.curselection()
        if select:
            varde = self.listbox_in.get(select)
            wb = openpyxl.load_workbook('Docs/Konton.xlsx', data_only=True)
            ws = wb["Konton"]

            output_lista = []
            for row in ws['B1:B1000']:
                for cell in row:
                    if cell.value != None:
                        kontogrupp = cell.value
                        if varde == kontogrupp:
                            kontonr = cell.offset(column=1).value
                            kontonamn = cell.offset(column=2).value
                            if kontonamn != None and len(str(kontonr)) > 3:
                                output = str(kontonr) + " " + str(kontonamn)
                                output_lista.append(output)
            wb.close()
            if output_lista:
                for x in output_lista:
                    self.listbox_kontogrupp.insert(END, x)


    def for_over_listbox_grupp(self):
        self.listbox_ut.insert(END, self.listbox_in.get(ANCHOR))
        self.lista_valda_kostnader.append(self.listbox_in.get(ANCHOR))

    def for_over_listbox_spec(self):
        self.listbox_ut.insert(END, self.listbox_kontogrupp.get(ANCHOR))
        self.lista_valda_kostnader.append(self.listbox_kontogrupp.get(ANCHOR))

    def ta_bort_listbox(self):
        self.lista_valda_kostnader.remove(self.listbox_ut.get(ANCHOR))
        self.listbox_ut.delete(ANCHOR)

    def spara_till_db(self):
        self.namn = self.finnamn.get()
        self.finnamn.delete(0, END)
        self.motpart = self.motp.get()
        self.motp.delete(0, END)
        self.per_f = self.perf.get()
        self.perf.delete(0, END)
        self.per_s = self.pers.get()
        self.pers.delete(0, END)
        self.radio_varde = self.radio_var.get()
        self.radio_var.set(0)
        self.ej_godk_kostnader = self.lista_valda_kostnader

        if self.ent_c > 0:
            self.procent_oh = self.oh_ent.get()
            self.ta_bort_entry()

        self.listbox_ut.delete(0, END)

        wb = openpyxl.load_workbook('Docs/Finansiarer.xlsx', data_only=True)
        ws = wb["Data"]

        ws.cell(row=ws.max_row + 1, column=1).value = self.namn
        ws.cell(row=ws.max_row, column=2).value = self.motpart
        ws.cell(row=ws.max_row, column=3).value = self.per_f
        ws.cell(row=ws.max_row, column=4).value = self.per_s

        if self.radio_varde == 0 or self.radio_varde == None:
            pass
        if self.radio_varde == 1:
            ws.cell(row=ws.max_row, column=5).value = "All OH godkänd"
        if self.radio_varde == 2:
            ws.cell(row=ws.max_row, column=5).value = "OH på tot. kostnader"
            ws.cell(row=ws.max_row, column=6).value = self.procent_oh
        if self.radio_varde == 3:
            ws.cell(row=ws.max_row, column=5).value = "OH på lön"
            ws.cell(row=ws.max_row, column=6).value = self.procent_oh
        if self.radio_varde == 4:
            ws.cell(row=ws.max_row, column=5).value = "Ingen OH"

        col = 7
        if self.ej_godk_kostnader:
            for x in self.ej_godk_kostnader:
                ws.cell(row=ws.max_row, column=col).value = x.split()[0]
                col+=1


        wb.save('Docs/Finansiarer.xlsx')
        wb.close()
        self.uppdatera_trad()
        self.tab3.uppdatera_droplist_finans()

    def initiera_trad(self):
        wb = openpyxl.load_workbook('Docs\\Finansiarer.xlsx', data_only=True)
        ws = wb['Data']
        self.lista_trad = []
        for row in ws['A2:A1000']:
            for cell in row:
                namn = cell.value
                motp = cell.offset(column=1).value
                per_f = cell.offset(column=2).value
                per_s = cell.offset(column=3).value
                oh = cell.offset(column=4).value
                oh_proc = cell.offset(column=5).value
                if namn != None:
                    lista_ej_godk = []
                    col = 6
                    while col < 20:
                        ej_godk = cell.offset(column=col).value
                        col += 1
                        if ej_godk != None:
                            lista_ej_godk.append(ej_godk)
                    ej_g = ', '.join(lista_ej_godk)
                    self.lista_trad.append([namn, motp, per_f, per_s, oh, oh_proc, ej_g])
        self.lista_valda_kostnader = []
        wb.close()

        c = 0
        for x in self.lista_trad:
            if x[5] != None:
                self.tree.insert(parent='', index='end', iid=c, values=(x[0], x[1], x[2], x[3], x[4]+", "+x[5]+"%", x[6]))
                c += 1
            else:
                self.tree.insert(parent='', index='end', iid=c, values=(x[0], x[1], x[2], x[3], x[4], x[6]))
                c += 1

    def uppdatera_trad(self):
        # Cleara Treeview
        c = 0
        for x in self.lista_trad:
            self.tree.delete(c)
            c += 1
        # Populera ny lista från databas
        self.initiera_trad()

    def ta_bort_fran_db(self):
        wb = openpyxl.load_workbook('Docs/Finansiarer.xlsx', data_only=True)
        ws = wb["Data"]

        # Hämta värden projektnummer från val i Treeview
        i = self.tree.focus()
        d = self.tree.item(i)
        val = d['values']
        namn = val[0]

        c = 1
        for row in ws['A2:A1000']:
            for cell in row:
                c += 1
                if str(cell.value) == str(namn):
                    ws.delete_rows(c)
        wb.save('Docs/Finansiarer.xlsx')
        wb.close()
        self.uppdatera_trad()
        self.tab3.uppdatera_droplist_finans()

















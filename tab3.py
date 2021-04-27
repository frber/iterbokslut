from tkinter import *
from tkinter.ttk import *
from ttkthemes import ThemedTk
from tkinter import filedialog
from tkinter import messagebox
import pandas as pd
import openpyxl
import os
import threading


from tab1 import *
from tab4 import *
from skapaperforslag import *


class Tab3:

    def __init__(self, tab1, tab3, tab4):

        self.tab1 = tab1
        self.tab3 = tab3
        self.tab4 = tab4

        rel_y = 40
        rel_x = 40

        # Projektnummer
        self.projnum_label = Label(self.tab3, text="Projektnummer:")
        self.projnum_label.place(y=rel_y, x=rel_x)
        self.projnum = Entry(self.tab3, width=25)
        self.projnum.place(y=rel_y, x=rel_x+100)


        # Projektnamn
        self.projnamn_label = Label(self.tab3, text="Projektnamn:")
        self.projnamn_label.place(y=rel_y+30, x=rel_x)
        self.projnamn = Entry(self.tab3, width=30)
        self.projnamn.place(y=rel_y+30, x=rel_x+100)

        # Droplist finansiär 1
        self.fin_label1 = Label(self.tab3, text="Finansiär 1")
        self.fin_label1.place(y=rel_y + 90, x=rel_x)
        self.finansiarer = self.hamta_fin()
        self.fin1 = StringVar()
        self.drop1 = OptionMenu(self.tab3, self.fin1, "", *self.finansiarer)
        self.drop1.place(y=rel_y + 120, x=rel_x)


        # Finansieringsgrad 1
        self.fingrad_label1 = Label(self.tab3, text="Finansieringsgrad:")
        self.fingrad_label1.place(y=rel_y+150, x=rel_x)
        self.fingrad1 = Entry(self.tab3, width=10)
        self.fingrad1.place(y=rel_y+150, x=rel_x+100)

        # Droplist finansiär 2
        self.fin_label2 = Label(self.tab3, text="Finansiär 2")
        self.fin_label2.place(y=rel_y + 200, x=rel_x)
        self.finansiarer = self.hamta_fin()
        self.fin2 = StringVar()
        self.drop2 = OptionMenu(self.tab3, self.fin2, "", *self.finansiarer)
        self.drop2.place(y=rel_y + 230, x=rel_x)

        # Finansieringsgrad 2
        self.fingrad_label2 = Label(self.tab3, text="Finansieringsgrad:")
        self.fingrad_label2.place(y=rel_y + 260, x=rel_x)
        self.fingrad2 = Entry(self.tab3, width=10)
        self.fingrad2.place(y=rel_y + 260, x=rel_x + 100)

        # Droplist finansiär 3
        self.fin_label3 = Label(self.tab3, text="Finansiär 3")
        self.fin_label3.place(y=rel_y + 310, x=rel_x)
        self.finansiarer = self.hamta_fin()
        self.fin3 = StringVar()
        self.drop3 = OptionMenu(self.tab3, self.fin3, "", *self.finansiarer)
        self.drop3.place(y=rel_y + 340, x=rel_x)

        # Finansieringsgrad 3
        self.fingrad_label3 = Label(self.tab3, text="Finansieringsgrad:")
        self.fingrad_label3.place(y=rel_y + 370, x=rel_x)
        self.fingrad3 = Entry(self.tab3, width=10)
        self.fingrad3.place(y=rel_y + 370, x=rel_x + 100)



        # Droplist projekt från db -tab3
        # self.lista_projekt_i_db = self.hamta_projekt()
        # self.proj_db = StringVar()
        # self.drop2 = OptionMenu(self.tab3, self.proj_db, "", *self.lista_projekt_i_db)
        # self.drop2.grid(row=6, column=1, sticky="W")

        # Knapp Lägg till -tab3
        self.knapp_lagg_till = Button(self.tab3, text="Lägg till", command=self.spara_till_db)
        self.knapp_lagg_till.place(y=280, x=400)

        # Knapp Ta bort -tab3
        self.knapp_ta_bort = Button(self.tab3, text="Ta bort", command=self.ta_bort_fran_db)
        self.knapp_ta_bort.place(y=280, x=1340)

        # Träd
        self.tree = Treeview(self.tab3)
        self.tree['columns'] = ("Projektnummer", "Projektnamn", "Finansiär 1", "Finansiär 2", "Finansiär 3")
        self.tree.column("#0", width=0, stretch=NO)
        self.tree.column("Projektnummer", anchor=W)
        self.tree.column("Projektnamn", anchor=W)
        self.tree.column("Finansiär 1", anchor=W)
        self.tree.column("Finansiär 2", anchor=W)
        self.tree.column("Finansiär 3", anchor=W)

        self.tree.heading("#0", text="", anchor=W)
        self.tree.heading("Projektnummer", text="Projektnummer", anchor=W)
        self.tree.heading("Projektnamn", text="Projektnamn", anchor=W)
        self.tree.heading("Finansiär 1", text="Finansiär 1", anchor=W)
        self.tree.heading("Finansiär 2", text="Finansiär 2", anchor=W)
        self.tree.heading("Finansiär 3", text="Finansiär 3", anchor=W)

        self.tree.place(y=50, x=400)


        self.initiera_trad()

        # Prognar -tab3
        # self.s = ttk.Style()
        # self.s.theme_use("winnative")
        # self.s.configure("blue.Horizontal.TProgressbar", foreground='navy', background='navy')

        #self.prog_bar.grid(row=10, column=5)

        self.placering_y = 600
        self.placering_x = 400
        #self.boxlist = []
        #elf.uppdatera_boxlista()
        self.tab4.uppdatera_boxlista()

        # Knapp Skapa perförslag -tab3

        #self.knapp_skapa_forslag.grid(row=10, column=4)



    def valj_gamla_berpers(self):
        self.filvag_gamla_berpers = StringVar()
        vald_filvag = filedialog.askdirectory()
        self.filvag_gamla_berpers.set(vald_filvag)
        self.label_gamla_berpers["text"] = vald_filvag

    def valj_spara_berpers(self):
        self.filvag_spara_berpers = StringVar()
        vald_filvag = filedialog.askdirectory()
        self.filvag_spara_berpers.set(vald_filvag)
        self.label_spara_berpers["text"] = vald_filvag

    def uppdatera_droplist_finans(self):
        self.finansiarer = self.hamta_fin()
        self.drop1.set_menu("", *self.finansiarer)
        self.fin1.set("")
        self.drop2.set_menu("", *self.finansiarer)
        self.fin2.set("")
        self.drop3.set_menu("", *self.finansiarer)
        self.fin3.set("")

    def hamta_fin(self):
        df = pd.read_excel(r'Docs\Finansiarer.xlsx')
        fin = df['FINANSIÄR'].tolist()
        return fin

    def hamta_projekt(self):
        wb = openpyxl.load_workbook('Docs/Projekt.xlsx', data_only=True)
        ws = wb["Projekt"]
        lista_projekt_i_db = []
        for row in ws['A2:A1000']:
            for cell in row:
                projnum = cell.value
                projnamn = cell.offset(column=1).value
                if projnum != None:
                    if projnamn != None:
                        lista_projekt_i_db.append(str(projnum) + " " + str(projnamn))
                    else:
                        lista_projekt_i_db.append(str(projnum))
        wb.close()
        return lista_projekt_i_db

    def spara_till_db(self):
        projektnummer = self.projnum.get()
        print(projektnummer)
        self.projnum.delete(0, END)
        projektnamn = self.projnamn.get()
        self.projnamn.delete(0, END)
        finansiar1 = self.fin1.get()
        self.fin1.set("")
        finansieringsgrad1 = self.fingrad1.get()
        self.fingrad1.delete(0, END)
        finansiar2 = self.fin2.get()
        self.fin2.set("")
        finansieringsgrad2 = self.fingrad2.get()
        self.fingrad2.delete(0, END)
        finansiar3 = self.fin3.get()
        self.fin3.set("")
        finansieringsgrad3 = self.fingrad3.get()
        self.fingrad3.delete(0, END)

        if self.kontrollera_dubbel(projektnummer):
            wb = openpyxl.load_workbook('Docs/Projekt.xlsx', data_only=True)
            ws = wb["Projekt"]
            ws.cell(row=ws.max_row + 1, column=1).value = projektnummer
            ws.cell(row=ws.max_row, column=2).value = projektnamn
            ws.cell(row=ws.max_row, column=3).value = finansiar1
            ws.cell(row=ws.max_row, column=4).value = finansieringsgrad1
            ws.cell(row=ws.max_row, column=5).value = finansiar2
            ws.cell(row=ws.max_row, column=6).value = finansieringsgrad2
            ws.cell(row=ws.max_row, column=7).value = finansiar3
            ws.cell(row=ws.max_row, column=8).value = finansieringsgrad3
            wb.save('Docs/Projekt.xlsx')
            wb.close()
            self.tab4.uppdatera_boxlista()
            self.uppdatera_trad_projekt()

    def kontrollera_dubbel(self, projektnummer):
        wb = openpyxl.load_workbook('Docs/Projekt.xlsx', data_only=True)
        ws = wb["Projekt"]
        kontroll_projekt = []
        for row in ws['A1:A1000']:
            for cell in row:
                if cell.value != None:
                    kontroll_projekt.append(str(cell.value))
        if str(projektnummer) in kontroll_projekt:
            messagebox.showerror("OBS!", "Projektnummer finns redan i databasen!")
        if len(projektnummer) == 0:
            messagebox.showerror("OBS!", "Skriv in ett projektnummer!")
        else:
            return True



    def uppdatera_droplist_projekt(self):
        # Används ej just nu
        self.lista_projekt_i_db = self.hamta_projekt()
        self.drop2.set_menu("", *self.lista_projekt_i_db)
        self.proj_db.set("")

    def ta_bort_fran_db(self):
        wb = openpyxl.load_workbook('Docs/Projekt.xlsx', data_only=True)
        ws = wb["Projekt"]

        # Hämta värden projektnummer från val i Treeview
        i = self.tree.focus()
        d = self.tree.item(i)
        val = d['values']
        projnr = val[0]

        c = 1
        for row in ws['A2:A1000']:
            for cell in row:
                c += 1
                if str(cell.value) == str(projnr):
                    ws.delete_rows(c)
        wb.save('Docs/Projekt.xlsx')
        wb.close()

        # self.uppdatera_droplist_projekt()
        #self.ta_bort_boxar()
        self.tab4.ta_bort_boxar()
        self.uppdatera_trad_projekt()

    def initiera_trad(self):
        wb = openpyxl.load_workbook('Docs\\Projekt.xlsx', data_only=True)
        ws = wb['Projekt']
        self.lista_trad = []
        for row in ws['A2:A1000']:
            for cell in row:
                projnr = cell.value
                projnamn = cell.offset(column=1).value
                fin1 = cell.offset(column=2).value
                fingrad1 = cell.offset(column=3).value
                fin2 = cell.offset(column=4).value
                fingrad2 = cell.offset(column=5).value
                fin3 = cell.offset(column=6).value
                fingrad3 = cell.offset(column=7).value
                if projnr != None:
                    if fin1 == None:
                        self.lista_trad.append([projnr, projnamn, "","",""])
                    if fin1 != None and fin2 == None:
                        self.lista_trad.append([projnr, projnamn, str(fin1)+", "+str(fingrad1)+"%", "", ""])
                    if fin1 != None and fin2 != None and fin3 == None:
                        self.lista_trad.append([projnr, projnamn, str(fin1) + ", " + str(fingrad1) + "%",
                        str(fin2) + ", " + str(fingrad2) + "%",
                        ""])
                    if fin1 != None and fin2 != None and fin3 != None:
                        self.lista_trad.append([projnr, projnamn, str(fin1) + ", " + str(fingrad1) + "%",
                                                str(fin2) + ", " + str(fingrad2) + "%",
                                                str(fin3) + ", " + str(fingrad3) + "%"])



        wb.close()

        c = 0
        for x in self.lista_trad:
            self.tree.insert(parent='', index='end', iid=c, values=(x[0], x[1], x[2], x[3], x[4]))
            c +=1

    def uppdatera_trad_projekt(self):
        # Cleara Treeview
        c = 0
        for x in self.lista_trad:
            self.tree.delete(c)
            c += 1
        # Populera ny lista från databas
        self.initiera_trad()








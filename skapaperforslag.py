import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import os
import shutil
from datetime import datetime



class SkapaPerForslag:

    def __init__(self, projnr_db, projnamn_db, lista_finansiarer, lista_fingrader, filvag_gamla_berpers, filvag_spara_berpers):
        self.projnr_db = projnr_db
        self.projnamn_db = projnamn_db
        self.lista_finansiarer = lista_finansiarer
        self.lista_fingrader = lista_fingrader
        self.filvag_gamla_berpers = filvag_gamla_berpers
        self.filvag_spara_berpers = filvag_spara_berpers
        self.avgor_vilket_bokslut()
        self.hamta_agressodata()

    def avgor_vilket_bokslut(self):
        self.idag = datetime.now()
        self.ar = self.idag.strftime("%Y")
        self.manad = self.idag.strftime("%m")
        self.manad = int(self.manad.split("0")[1])

        if self.manad > 2 and self.manad <= 5:
            self.bokslutperiod = "T1"
        if 5 < self.manad < 12:
            self.bokslutperiod = "T2"
        if self.manad < 2 or self.manad == 12:
            self.bokslutperiod = "T3"


    def hamta_agressodata(self):
        wb_agresso = openpyxl.load_workbook('Docs/Agressodata.xlsx', data_only=True)
        ws_agresso = wb_agresso['Agressodata']
        self.agressodata = []
        c = 0
        for row in ws_agresso['C1:C1000']:
            for cell in row:
                projektnr = cell.value
                konto = cell.offset(column=-2).value
                konto_text = cell.offset(column=-1).value
                #projektnr = cell.offset(column=2).value
                vht = cell.offset(column=2).value
                motp = cell.offset(column=3).value
                finans = cell.offset(column=4).value
                belopp = cell.offset(column=5).value

                if str(projektnr) == str(self.projnr_db):
                    c += 1
                    self.agressodata.append([konto, konto_text, projektnr, vht, motp, finans, belopp])
        wb_agresso.close()
        if c > 0:
            self.leta_gamal_berper()
            #if not self.leta_gamal_berper():
                #self.skapa_berper()

    def leta_gamal_berper(self):
        for root, dirs, files in os.walk(self.filvag_gamla_berpers):
            for fil in files:
                filvag_gamal_berper = os.path.join(root, fil).replace("\\", "/")
                if filvag_gamal_berper.endswith(".xlsx") or filvag_gamal_berper.endswith(".xlsm"):
                    storlek = os.path.getsize(filvag_gamal_berper)
                    if storlek < 700000:
                        try:
                            wb_gammal = openpyxl.load_workbook(filvag_gamal_berper, data_only=True)
                        except:
                            self.skapa_berper()
                        else:
                            for sheet in wb_gammal.worksheets:
                                projnummer_gammal = sheet.cell(32, 8).value
                                if str(projnummer_gammal) == str(self.projnr_db):
                                    ny_berper = self.filvag_spara_berpers + "\\" + str(self.projnr_db) + ".xlsx"
                                    shutil.copy(filvag_gamal_berper, ny_berper)
                                    wb_berper = openpyxl.load_workbook(ny_berper, data_only=True)
                                    ny_sheet = 'Periodiseringsförslag '+str(self.bokslutperiod)+ " "+str(self.ar)
                                    wb_berper.create_sheet(ny_sheet)
                                    wb_berper.save(ny_berper)
                                    self.for_over_agressodata(wb_berper, ny_berper, ny_sheet)
                                    break
                    else:
                        self.skapa_berper()


                # NAMNALTERNATIV----------------
                #filvag_gamal_berper = os.path.join(root, fil).replace("\\","/")
                #bara_filnamn = str(fil.split(".")[0])
                #projnr_filnamn = bara_filnamn.split(" ")[0]
                #if projnr_filnamn == self.projnr_db:
                    #ny_berper = self.filvag_spara_berpers+"\\"+str(self.projnr_db)+".xlsx"
                    #shutil.copy(filvag_gamal_berper, ny_berper)
                    #try:
                        #wb = openpyxl.load_workbook(ny_berper, data_only=True)
                    #except:
                        #self.skapa_berper()
                        #continue
                    #ny_sheet = 'Periodiseringsförslag '+str(self.bokslutperiod)+ " "+str(self.ar)
                    #print(ny_sheet)
                    #wb.create_sheet(ny_sheet)
                    #wb.save(ny_berper)
                    #self.for_over_agressodata(wb, ny_berper, ny_sheet)
                    #return True


    def skapa_berper(self):
        orginal_berper = 'Docs\\Berper.xlsx'
        ny_berper =  self.filvag_spara_berpers+"\\"+str(self.projnr_db)+".xlsx"
        shutil.copy(orginal_berper, ny_berper)
        wb_berper = openpyxl.load_workbook(ny_berper, data_only=True)
        ny_sheet = 'Periodiseringsförslag ' + str(self.bokslutperiod) + " " + str(self.ar)
        wb_berper.create_sheet(ny_sheet)
        wb_berper.save(ny_berper)
        self.for_over_agressodata(wb_berper, ny_berper, ny_sheet)


    def for_over_agressodata(self, wb_berper, ny_berper, ny_sheet):
        ws_berper_perforslag = wb_berper[ny_sheet]
        for x in self.agressodata:
            konto = x[0]
            kontotext = x[1]
            projektnr = x[2]
            vht = x[3]
            motp = x[4]
            fin = x[5]
            belopp = x[6]

            ws_berper_perforslag.column_dimensions['A'].width = 10
            ws_berper_perforslag.column_dimensions['B'].width = 50
            ws_berper_perforslag.column_dimensions['C'].width = 10
            ws_berper_perforslag.column_dimensions['D'].width = 10
            ws_berper_perforslag.column_dimensions['E'].width = 10
            ws_berper_perforslag.column_dimensions['F'].width = 10
            ws_berper_perforslag.column_dimensions['G'].width = 25

            stilborder_bot = Border(bottom=Side(style='thin'))


            konto_rubrik = ws_berper_perforslag.cell(row=1, column=1)
            konto_rubrik.value = "Konto"
            konto_rubrik.alignment = Alignment(horizontal='right')
            konto_rubrik.border = stilborder_bot
            konto_rubrik.font = Font(bold=True)

            kontotext_rubrik = ws_berper_perforslag.cell(row=1, column=2)
            kontotext_rubrik.value = "Kontotext"
            kontotext_rubrik.alignment = Alignment(horizontal='right')
            kontotext_rubrik.border = stilborder_bot
            kontotext_rubrik.font = Font(bold=True)

            projekt_rubrik = ws_berper_perforslag.cell(row=1, column=3)
            projekt_rubrik.value = "Projekt"
            projekt_rubrik.alignment = Alignment(horizontal='right')
            projekt_rubrik.border = stilborder_bot
            projekt_rubrik.font = Font(bold=True)

            vht_rubrik = ws_berper_perforslag.cell(row=1, column=4)
            vht_rubrik.value = "Vht"
            vht_rubrik.alignment = Alignment(horizontal='right')
            vht_rubrik.border = stilborder_bot
            vht_rubrik.font = Font(bold=True)

            motp_rubrik = ws_berper_perforslag.cell(row=1, column=5)
            motp_rubrik.value = "Motpart"
            motp_rubrik.alignment = Alignment(horizontal='right')
            motp_rubrik.border = stilborder_bot
            motp_rubrik.font = Font(bold=True)

            fin_rubrik = ws_berper_perforslag.cell(row=1, column=6)
            fin_rubrik.value = "Finansiär"
            fin_rubrik.alignment = Alignment(horizontal='right')
            fin_rubrik.border = stilborder_bot
            fin_rubrik.font = Font(bold=True)

            belopp_rubrik = ws_berper_perforslag.cell(row=1, column=7)
            belopp_rubrik.value = "Belopp"
            belopp_rubrik.alignment = Alignment(horizontal='right')
            belopp_rubrik.border = stilborder_bot
            belopp_rubrik.font = Font(bold=True)



            cell_konto =  ws_berper_perforslag.cell(row=ws_berper_perforslag.max_row+1, column=1)
            cell_konto.value = konto
            cell_kontotext = ws_berper_perforslag.cell(row=ws_berper_perforslag.max_row, column=2)
            cell_kontotext.value = kontotext
            cell_projektnummer = ws_berper_perforslag.cell(row=ws_berper_perforslag.max_row, column=3)
            cell_projektnummer.value = projektnr
            cell_vht = ws_berper_perforslag.cell(row=ws_berper_perforslag.max_row, column=4)
            cell_vht.value = vht
            cell_motp = ws_berper_perforslag.cell(row=ws_berper_perforslag.max_row, column=5)
            cell_motp.value = motp
            cell_fin = ws_berper_perforslag.cell(row=ws_berper_perforslag.max_row, column=6)
            cell_fin.value = fin
            cell_belopp = ws_berper_perforslag.cell(row=ws_berper_perforslag.max_row, column=7)
            cell_belopp.number_format = '#,##0.00'
            cell_belopp.value = belopp
            if belopp < 0:
                cell_belopp.font = Font(color='00FF0000')
            if konto == "B" or konto == "D":
                cell_konto.font = Font(bold=True)
                cell_kontotext.font = Font(bold=True)
                cell_projektnummer.font = Font(bold=True)
                cell_belopp.font = Font(bold=True)

                stilborder = Border(top=Side(style='thin'), bottom=Side(style='thin'))
                cell_konto.border = stilborder
                cell_kontotext.border = stilborder
                cell_projektnummer.border = stilborder
                cell_vht.border = stilborder
                cell_motp.border = stilborder
                cell_fin.border = stilborder
                cell_belopp.border = stilborder

                if belopp < 0:
                    cell_belopp.font = Font(color='00FF0000', bold=True)

            if konto == None and kontotext == None and projektnr != None and belopp != None:
                cell_kontotext.value = "Totalt resultat"
                cell_projektnummer.font = Font(bold=True)
                cell_belopp.font = Font(bold=True)
                cell_kontotext.font = Font(bold=True)
                cell_konto.border = stilborder
                cell_kontotext.border = stilborder
                cell_projektnummer.border = stilborder
                cell_vht.border = stilborder
                cell_motp.border = stilborder
                cell_fin.border = stilborder
                cell_belopp.border = stilborder
                if belopp < 0:
                    cell_belopp.font = Font(color='00FF0000', bold=True)





        self.ratt_bokslutsperiod(wb_berper)
        wb_berper.save(ny_berper)
        wb_berper.close()
        self.kalk_periodiseringsforslag(wb_berper, ws_berper_perforslag, ny_berper)


    def ratt_bokslutsperiod(self, wb_berper):
        # Fyller i rätt bokslutsperiod i förstasidan för berper.
        if self.bokslutperiod == "T1":
            ratt_bokslutsperiod = int(self.ar+"04")
        if self.bokslutperiod == "T2":
            ratt_bokslutsperiod = int(self.ar+"08")
        if self.bokslutperiod == "T3":
            ratt_bokslutsperiod = int(self.ar+"12")

        for sheet in wb_berper.worksheets:
            hook = sheet.cell(31, 10).value
            hook = str(hook)
            hook = hook.lower()
            if hook == "bokslutsperiod":
                datum = sheet.cell(31, 11)
                datum.value = ratt_bokslutsperiod

    def kalk_periodiseringsforslag(self, wb_berper, ws_berper_perforslag, ny_berper):

        #print("hej")

        #Rubrik direkta kostnader, lönekostnader, periodisering, bidrag
        bold = Font(bold=True)
        ws_berper_perforslag.column_dimensions['H'].width = 10
        ws_berper_perforslag.column_dimensions['I'].width = 25
        ws_berper_perforslag.column_dimensions['J'].width = 15


        cell_direkta_kostnader_rubrik = ws_berper_perforslag.cell(row=2, column=9)
        cell_direkta_kostnader_rubrik.value = "Totala direkta kostnader"
        cell_direkta_kostnader_rubrik.font = bold
        cell_lonekostander_rubrik = ws_berper_perforslag.cell(row=3, column=9)
        cell_lonekostander_rubrik.value = "Totala Lönekostnader"
        cell_lonekostander_rubrik.font = bold
        #cell_periodisering_rubrik = ws_berper_perforslag.cell(row=4, column=7)
        #cell_periodisering_rubrik.value = "Tidigare periodisering"
        #cell_periodisering_rubrik.font = bold
        #cell_bidrag_rubrik = ws_berper_perforslag.cell(row=5, column=7)
        #cell_bidrag_rubrik.value = "Tidigare erhållet bidrag"
        #cell_bidrag_rubrik.font = bold


        cell_direkta_kostnader_varde = ws_berper_perforslag.cell(row=2, column=10)
        cell_direkta_kostnader_varde.value = "=0"
        cell_direkta_kostnader_varde.number_format = '#,##0.00'
        cell_lonekostander_varde = ws_berper_perforslag.cell(row=3, column=10)
        cell_lonekostander_varde.value = "=0"
        cell_lonekostander_varde.number_format = '#,##0.00'








        lista_direkta_kostnader = self.hamta_direkta_och_lonekostnader()[0]
        lista_lonekostnader = self.hamta_direkta_och_lonekostnader()[1]




        for row in ws_berper_perforslag['A1:A1000']:
            for cell in row:
                if cell.value != None:
                    kontonr_berper = cell.value
                    kontonamn_berper = cell.offset(column=1).value
                    belopp_cell = cell.offset(column=6)

                    # Avgör ej godkända kostnader
                    #for x in lista_ej_godk_kost_utokad:
                        #if str(kontonr_berper) == str(x):
                            #ej_godk_kostnader = ws_berper_perforslag.cell(rad_kost, col)
                            # ej_godk_kostnader.value = "="+str(belopp_cell.coordinate)
                            #rad_kost += 1
                    # Avgör godkända kostnader
                    #if isinstance(kontonr_berper, int):
                        #if str(kontonr_berper) != str(x) and str(kontonr_berper)[0] != "3":
                            #godk_kostnader = ws_berper_perforslag.cell(rad_kost2, col + 1)
                            # godk_kostnader.value = "="+str(belopp_cell.coordinate)
                            #rad_kost2 += 1

                    # Avgör direkta kostnader
                    # c2 = 0
                    for y in lista_direkta_kostnader:
                        if str(kontonr_berper) == str(y):
                            cell_direkta_kostnader_varde.value += "+" + belopp_cell.coordinate

                    # Avgör lönekostnader
                    for z in lista_lonekostnader:
                        if str(kontonr_berper) == str(z):
                            cell_lonekostander_varde.value += "+" + belopp_cell.coordinate

        rad = 2
        col = 9
        fin_c = 0

        ws_berper_perforslag.column_dimensions['K'].width = 7

        ws_berper_perforslag.column_dimensions['L'].width = 20
        ws_berper_perforslag.column_dimensions['M'].width = 12

        ws_berper_perforslag.column_dimensions['N'].width = 7

        ws_berper_perforslag.column_dimensions['O'].width = 20
        ws_berper_perforslag.column_dimensions['P'].width = 12

        ws_berper_perforslag.column_dimensions['Q'].width = 7

        ws_berper_perforslag.column_dimensions['R'].width = 20
        ws_berper_perforslag.column_dimensions['S'].width = 12

        for finansiar, fingrad in zip(self.lista_finansiarer, self.lista_fingrader):
            if finansiar != None:
                rad = 2
                col += 3
                motpart = self.hamta_info_fin(finansiar)[0]
                per_fordran = self.hamta_info_fin(finansiar)[1]
                per_skuld = self.hamta_info_fin(finansiar)[2]
                oh = self.hamta_info_fin(finansiar)[3]
                oh_procent = self.hamta_info_fin(finansiar)[4]
                lista_ej_godk_kost = self.hamta_info_fin(finansiar)[5]
                lista_ej_godk_kost_utokad = self.hamta_lista_ej_godk(lista_ej_godk_kost)



                fin_c += 1
                cell_fin_rubr = ws_berper_perforslag.cell(rad, col)
                cell_fin_rubr.value = "Finansiar"+" "+str(fin_c)+":"
                cell_fin_rubr.alignment = Alignment(horizontal='center')

                cell_fin_varde = ws_berper_perforslag.cell(rad, col+1)
                cell_fin_varde.value = finansiar
                cell_fin_varde.alignment = Alignment(horizontal='center')

                cell_fingrad_rubr = ws_berper_perforslag.cell(rad+1, col)
                cell_fingrad_rubr.value = "Finanseringsgrad:"
                cell_fingrad_rubr.alignment = Alignment(horizontal='center')

                cell_fingrad_varde = ws_berper_perforslag.cell(rad+1, col+1)
                cell_fingrad_varde.alignment = Alignment(horizontal='center')
                cell_fingrad_varde.number_format = '0%'
                cell_fingrad_varde.value = int(fingrad)/100

                cell_oh_rubr = ws_berper_perforslag.cell(rad+2, col)
                cell_oh_rubr.value = "Godkänd OH:"
                cell_oh_rubr.alignment = Alignment(horizontal='center')

                cell_oh_varde = ws_berper_perforslag.cell(rad+2, col+1)
                cell_oh_varde.value = "=0"
                cell_oh_varde.number_format = '#,##0.00'
                cell_oh_varde.alignment = Alignment(horizontal='center')

                if oh == "All OH godkänd":
                    for row in ws_berper_perforslag['A1:A1000']:
                        for cell in row:
                            if cell.value != None:
                                kontonr_berper = cell.value
                                #print(kontonr_berper
                                belopp_cell = cell.offset(column=6)
                               # print(belopp_cell)
                                if kontonr_berper == 57990 and belopp_cell.value != None or kontonr_berper == 57991 and belopp_cell.value != None:
                                    print(belopp_cell.coordinate)
                                    cell_oh_varde.value += "+" + belopp_cell.coordinate









                cell_ej_godk = ws_berper_perforslag.cell(rad+1, col)
                #cell_ej_godk.value = "Ej godkända kostnader"
                #cell_godk = ws_berper_perforslag.cell(rad+1, col+1)
                #cell_godk.value = "Godkända kostnader"

                #rad_kost = 4
                #rad_kost2 = 4

                #col_direkta_lone_kostnader = 6
                #rad_direkta_lone_kostnader = 2












        wb_berper.save(ny_berper)

    def hamta_direkta_och_lonekostnader(self):
        wb_konton = openpyxl.load_workbook('Docs\\Konton.xlsx', data_only=True)
        ws_direkta = wb_konton['Direkta kostnader']
        ws_loner = wb_konton['Lönekostnader']

        lista_direkta = []
        lista_loner = []

        for row in ws_direkta['C1:C1000']:
            for cell in row:
                if cell.value != None:
                    lista_direkta.append(cell.value)

        for row in ws_loner['C1:C1000']:
            for cell in row:
                if cell.value != None:
                    lista_loner.append(cell.value)

        return lista_direkta, lista_loner



    def hamta_info_fin(self, finansiar):
        wb_fin = openpyxl.load_workbook('Docs\\Finansiarer.xlsx', data_only=True)
        ws_fin = wb_fin['Data']
        for row in ws_fin['A1:A1000']:
            for cell in row:
                if cell.value != None:
                    if cell.value == finansiar:
                        motpart = cell.offset(column=1).value
                        per_f = cell.offset(column=2).value
                        per_s = cell.offset(column=3).value
                        oh = cell.offset(column=4).value
                        oh_procent = cell.offset(column=5).value
                        lista_ej_godk = []
                        col = 6
                        while col < 20:
                            ej_godk = cell.offset(column=col).value
                            col += 1
                            if ej_godk != None:
                                lista_ej_godk.append(ej_godk)
                        return [motpart, per_f, per_s, oh, oh_procent, lista_ej_godk]

    def hamta_lista_ej_godk(self, lista_ej_godk_kost):
        wb_konton = openpyxl.load_workbook('Docs\\Konton.xlsx', data_only=True)
        ws_konton = wb_konton['Konton']
        lista_ej_godk_kost_utokad = []
        for x in lista_ej_godk_kost:
            for row in ws_konton['B1:B1000']:
                for cell in row:
                    kontonamn = cell.value
                    kontonummer = cell.offset(column=1).value
                    if str(kontonamn) == str(x) or str(kontonummer) == str(x):
                        lista_ej_godk_kost_utokad.append(kontonummer)

        return lista_ej_godk_kost_utokad















































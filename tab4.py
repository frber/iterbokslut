from tkinter import *
from tkinter.ttk import *
from ttkthemes import ThemedTk
from tkinter import filedialog
from tkinter import messagebox
import pandas as pd
import openpyxl
import os
import threading


from tab1 import *
from skapaperforslag import *


class Tab4:

    def __init__(self, tab4, tab1):
        self.tab4 = tab4
        self.tab1 = tab1


        self.label_forslag =  Label(self.tab4, text="Möjliga projekt att skapa periodseringsförslag för")
        self.label_forslag.grid(row=0, column=0, pady=5, padx=10)
        self.label_saknade = Label(self.tab4, text="Projekt som finns i agressodata men inte i din projektdatabas")
        self.label_saknade.grid(row=0, column=4, pady=5, padx=10)
        self.labellist = []
        self.skapa_labels_saknade_projekt()
        self.knapp_skapa_forslag = Button(self.tab4, text="Skapa Periodiseringsförslag", command=self.thread)
        self.knapp_skapa_forslag.grid(row=26,column=0)
        self.prog_bar = Progressbar(self.tab4, orient=HORIZONTAL, length=100, maximum=100, mode='indeterminate')
        self.prog_bar.grid(row=27, column=0)

        self.boxlist = []




    def skapa_labels_saknade_projekt(self):
        lista_proj_agresso = self.hamta_listor_agresso_projekt()[0]
        lista_proj = self.hamta_listor_agresso_projekt()[1]
        rensad_lista = []
        for x in lista_proj:
            rensad_lista.append(x.split()[0])

        rensad_lista = set(rensad_lista)
        rensad_lista = list(rensad_lista)
        lista_proj_agresso = set(lista_proj_agresso)
        lista_proj_agresso = list(lista_proj_agresso)

        rad = 1
        for y in lista_proj_agresso:
            if y not in rensad_lista and y.isdigit():
                label_proj =  Label(self.tab4, text=y)
                label_proj.grid(row=rad, column=4)
                self.labellist.append(label_proj)
                rad +=1


    def reset_label(self):
        if self.labellist:
            for x in self.labellist:
                x.destroy()
            self.skapa_labels_saknade_projekt()






    def uppdatera_boxlista(self):

        lista_proj_agresso = self.hamta_listor_agresso_projekt()[0]
        lista_proj = self.hamta_listor_agresso_projekt()[1]

        self.boxlist_utfall = []
        if lista_proj_agresso and lista_proj:
            lista_proj_agresso = set(lista_proj_agresso)
            rad = 1
            rad2 = 1
            rad3 = 1
            for x in lista_proj:
                if str(x.split()[0]) in lista_proj_agresso:
                    box = IntVar()
                    checkbox = Checkbutton(self.tab4, text=x, variable=box)
                    checkbox.grid(row=rad, column=0, sticky="W", pady=2, padx=10)
                    rad += 1
                    if rad > 25:
                        checkbox.grid(row=rad2, column=1, sticky="W", pady=2, padx=10)
                        rad2 += 1
                    if rad2 > 25:
                        checkbox.grid(row=rad3, column=2, sticky="W", pady=2, padx=10)
                        rad3 += 1

                    self.boxlist_utfall.append([box, x])
                    self.boxlist.append(checkbox)



    def hamta_listor_agresso_projekt(self):
        wb_agressodata = openpyxl.load_workbook('Docs/Agressodata.xlsx', data_only=True)
        ws_agresso = wb_agressodata['Agressodata']
        lista_proj_agresso = []
        for row in ws_agresso['C1:C1000']:
            for cell in row:
                if cell.value != None:
                    lista_proj_agresso.append(str(cell.value))

        wb_agressodata.close()

        wb = openpyxl.load_workbook('Docs/Projekt.xlsx', data_only=True)
        ws = wb["Projekt"]
        lista_proj = []
        for row in ws['A2:A1000']:
            for cell in row:
                projnr = cell.value
                projnamn = cell.offset(column=1).value
                if projnr != None:
                    if projnamn != None:
                        lista_proj.append(str(projnr) + " " + str(projnamn))
                    else:
                        lista_proj.append(str(projnr))
        wb.close()

        return lista_proj_agresso, lista_proj


    def ta_bort_boxar(self):
        for x in self.boxlist:
            x.destroy()
        self.uppdatera_boxlista()

    def thread(self):
        # Använder en annan thread så att gränssnittet inte fryser medans huvudprogrammet körs.
        # Startar lokalt i en egen metod eftersom detta måste instansieras på nytt, annars: RuntimeError: threads can only be started once.
        t = threading.Thread(target=self.skapa_perforslag, daemon=True)
        t.start()


    def skapa_perforslag(self):
        self.prog_bar.start(4)
        filvag_gamla_berpers = r'C:\Users\berfre\Desktop\gamla berperr'
        filvag_spara_berpers = r'C:\Users\berfre\Desktop\testspara'
        # Lägg till för dynamiskt sen
        # filvag_gamla_berpers = self.tab1.get_fivlag_gamla_berper()
        # filvag_spara_berpers = self.tab1.get_filvag_spara_berpers()
        wb = openpyxl.load_workbook('Docs/Projekt.xlsx', data_only=True)
        ws = wb["Projekt"]


        for x in self.boxlist_utfall:
            if x[0].get() == 1:
                projnr_box = x[1].split()[0]
                for row in ws['A2:A1000']:
                    for cell in row:
                        projnr_db = cell.value
                        projnamn_db = cell.offset(column=1).value
                        finansiar1_db = cell.offset(column=2).value
                        fingrad1_db = cell.offset(column=3).value
                        finansiar2_db = cell.offset(column=4).value
                        fingrad2_db = cell.offset(column=5).value
                        finansiar3_db = cell.offset(column=6).value
                        fingrad3_db = cell.offset(column=7).value


                        if projnr_box == projnr_db:
                            lista_finansiarer = [finansiar1_db, finansiar2_db, finansiar3_db]
                            lista_fingrader = [fingrad1_db, fingrad2_db, fingrad3_db]
                            skapa_per_forslag = SkapaPerForslag(projnr_db, projnamn_db, lista_finansiarer, lista_fingrader,
                                                                filvag_gamla_berpers, filvag_spara_berpers)
        wb.close()
        self.prog_bar.stop()

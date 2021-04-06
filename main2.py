from tkinter import *
from tkinter.ttk import *
import pandas as pd
import openpyxl




def spara_till_db(projnum, projnamn, fin, fingrad, tab2):
    projektnummer = projnum.get()
    projnum.delete(0, END)
    projektnamn = projnamn.get()
    projnamn.delete(0, END)
    finansiar = fin.get()
    fin.set("")
    finansieringsgrad = fingrad.get()
    fingrad.delete(0, END)
    wb = openpyxl.load_workbook('Docs/Projekt.xlsx', data_only=True)
    ws = wb["Projekt"]

    ws.cell(row=ws.max_row + 1, column=1).value = projektnummer
    ws.cell(row=ws.max_row, column=2).value = projektnamn
    ws.cell(row=ws.max_row, column=3).value = finansiar
    ws.cell(row=ws.max_row, column=4).value = finansieringsgrad
    wb.save('Docs/Projekt.xlsx')
    wb.close()
    uppdatera_db(tab2)

def uppdatera_db(tab2):
    wb = openpyxl.load_workbook('Docs/Projekt.xlsx', data_only=True)
    ws = wb["Projekt"]
    lista_projekt_i_db = []
    for row in ws['A2:A1000']:
        for cell in row:
            projnum = cell.value
            projnamn = cell.offset(column = 1).value
            if projnum != None:
                if projnamn != None:
                    lista_projekt_i_db.append(str(projnum)+" "+str(projnamn))
                else:
                    lista_projekt_i_db.append(str(projnum))
    proj_db = StringVar()
    drop2 = OptionMenu(tab2, proj_db, "", *lista_projekt_i_db).grid(row=6, column=1, sticky="W")







def skapa_falt(root, finansiarer):

    # Skapa tabs
    tabcontrol = Notebook(root)
    tab1 = Frame(tabcontrol)
    tabcontrol.add(tab1, text="Kontroll")
    #tabcontrol.pack(expand=1, fill="both")
    #ttk.Style().configure("TNotebook", bg="black")
    tab2 = Frame(tabcontrol)
    tabcontrol.add(tab2, text="Hantera dokument")
    tabcontrol.pack(expand = 1, fill ="both")

    # Projektnummer
    projnum = Entry(tab2, width=30)
    projnum.grid(row=1, column=1)
    projnum_label = Label(tab2, text="Projektnummer:").grid(row=1, column=0)

    # Projektnamn
    projnamn = Entry(tab2, width=30)
    projnamn.grid(row=2, column=1)
    projnamn_label = Label(tab2, text="Projektnamn:").grid(row=2, column=0)

    # Finansieringsgrad
    fingrad = Entry(tab2, width=30)
    fingrad.grid(row=4, column=1)
    fingrad_label = Label(tab2, text="Finansieringsgrad:").grid(row=4, column=0)

    # Droplist finansiärer
    fin = StringVar()
    drop = OptionMenu(tab2, fin, "", *finansiarer).grid(row=3, column=1, sticky="W")
    fin_label = Label(tab2, text="Finansiär:").grid(row=3, column=0)

    # Knapp Lägg till
    knapp_lagg_till = Button(tab2, text="Lägg till", command=lambda: spara_till_db(projnum, projnamn, fin, fingrad, tab2))
    knapp_lagg_till.grid(row=5, column=1)

    # Knapp Ta bot
    knapp_ta_bort = Button(tab2, text="Ta bort", command=lambda: spara_till_db(projnum, projnamn, fin, fingrad, tab2))
    knapp_ta_bort.grid(row=7, column=1)

    #  Skapa/uppdatera droplist för projekt sparat i databas
    uppdatera_db(tab2)

def hamta_fin():
    df = pd.read_excel(r'Docs\Data.xlsx')
    fin = df['FINANSIÄR'].tolist()
    return fin

def main():
    root = Tk()
    root.geometry("680x350")
    finansiarer = hamta_fin()
    skapa_falt(root, finansiarer)
    root.mainloop()

if __name__ == "__main__":
    main()